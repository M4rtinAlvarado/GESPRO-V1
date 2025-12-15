import os
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from copy import copy
from django.conf import settings
from django.db.models import Prefetch

from proyectos.models import (
    Proyecto, Actividad, ActividadDifusion, Actividad_Encargado,
    Periodo, ProductoAsociado, ActividadDifusion_Linea, EstadoActividad
)


def obtener_datos_proyecto(proyecto_id):
    """
    Obtiene todos los datos necesarios del proyecto para exportar.
    Retorna las actividades normales y de difusión con sus datos relacionados.
    """
    proyecto = Proyecto.objects.get(id=proyecto_id)
    
    # Prefetch para actividades normales
    encargados_prefetch = Prefetch(
        'actividad_encargados',
        queryset=Actividad_Encargado.objects.select_related('encargado')
            .only('encargado__nombre', 'encargado__correo_electronico'),
        to_attr='encargados_cache'
    )

    periodos_prefetch = Prefetch(
        'fechas',
        queryset=Periodo.objects.only('fecha_inicio', 'fecha_fin', 'estado'),
        to_attr='periodos_cache'
    )

    productos_prefetch = Prefetch(
        'productos_asociados',
        queryset=ProductoAsociado.objects.only('nombre'),
        to_attr='productos_cache'
    )

    # Obtener actividades normales del proyecto
    actividades_normales = (
        Actividad.objects
        .filter(linea_trabajo__proyecto=proyecto)
        .select_related('linea_trabajo')
        .prefetch_related(
            encargados_prefetch,
            periodos_prefetch,
            productos_prefetch
        )
        .only(
            'id',
            'n_act',
            'nombre',
            'linea_trabajo__nombre',
        )
        .order_by('n_act')
    )

    lista_actividades = []
    for act in actividades_normales:
        responsables = [
            f"{ae.encargado.nombre}, {ae.encargado.correo_electronico}"
            for ae in act.encargados_cache
        ]
        periodos = [(p.fecha_inicio, p.fecha_fin, p.estado) for p in act.periodos_cache]
        producto = act.productos_cache[0].nombre if act.productos_cache else ""

        fila = {
            'n_act': act.n_act,
            'nombre': act.nombre,
            'linea_trabajo': act.linea_trabajo.nombre,
            'responsables': responsables,
            'producto': producto,
            'periodos': periodos
        }
        lista_actividades.append(fila)

    # Prefetch para actividades de difusión
    lineas_prefetch = Prefetch(
        'actividad_lineas',
        queryset=ActividadDifusion_Linea.objects.select_related('linea_trabajo')
            .only('linea_trabajo__nombre'),
        to_attr='lineas_cache'
    )

    actividades_difusion = (
        ActividadDifusion.objects
        .filter(proyecto=proyecto)
        .prefetch_related(
            encargados_prefetch,
            periodos_prefetch,
            productos_prefetch,
            lineas_prefetch
        )
        .only(
            'id',
            'n_act',
            'nombre',
        )
        .order_by('n_act')
    )

    lista_actividades_difusion = []
    for act in actividades_difusion:
        responsables = [
            f"{ae.encargado.nombre}, {ae.encargado.correo_electronico}"
            for ae in act.encargados_cache
        ]
        periodos = [(p.fecha_inicio, p.fecha_fin, p.estado) for p in act.periodos_cache]
        productos = [p.nombre for p in act.productos_cache]
        lineas = [rel.linea_trabajo.nombre for rel in act.lineas_cache]

        fila = {
            'n_act': act.n_act,
            'nombre': act.nombre,
            'lineas_trabajo': lineas,
            'responsables': responsables,
            'productos': productos,
            'periodos': periodos
        }
        lista_actividades_difusion.append(fila)

    return proyecto, lista_actividades, lista_actividades_difusion


def obtener_rango_fechas(actividades, actividades_difusion):
    """
    Calcula el rango de fechas (mínima y máxima) de todas las actividades.
    """
    todas_fechas = []
    
    for act in actividades:
        for fecha_inicio, fecha_fin, _ in act['periodos']:
            todas_fechas.append(fecha_inicio)
            todas_fechas.append(fecha_fin)
    
    for act in actividades_difusion:
        for fecha_inicio, fecha_fin, _ in act['periodos']:
            todas_fechas.append(fecha_inicio)
            todas_fechas.append(fecha_fin)
    
    if not todas_fechas:
        # Si no hay fechas, usar el mes actual
        hoy = datetime.now().date()
        return hoy.replace(day=1), hoy.replace(day=28)
    
    return min(todas_fechas), max(todas_fechas)


def generar_columnas_fechas(ws, fecha_inicio, fecha_fin, col_inicio=7):
    """
    Genera las columnas de fechas en el encabezado (fila 1 y 2).
    Las fechas son siempre lunes de cada semana.
    Retorna un diccionario que mapea fechas a columnas.
    """
    fecha_columna = {}
    col_actual = col_inicio
    
    # Encontrar el primer lunes del mes de inicio
    primer_dia_mes = fecha_inicio.replace(day=1)
    # weekday(): 0=lunes, 1=martes, ..., 6=domingo
    dias_hasta_lunes = (7 - primer_dia_mes.weekday()) % 7
    if dias_hasta_lunes == 0 and primer_dia_mes.weekday() != 0:
        dias_hasta_lunes = 7
    # Si el primer día ya es lunes, usar ese
    if primer_dia_mes.weekday() == 0:
        fecha_actual = primer_dia_mes
    else:
        fecha_actual = primer_dia_mes + timedelta(days=dias_hasta_lunes)
    
    # Estilos para encabezado
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(bold=True, size=9)
    center_align = Alignment(horizontal='center', vertical='center')
    
    meses_espanol = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    mes_actual = None
    col_inicio_mes = col_actual
    columnas_por_mes = []  # Lista de (col_inicio, col_fin, mes)
    
    while fecha_actual <= fecha_fin + timedelta(days=7):
        # Escribir día en fila 2
        ws.cell(row=2, column=col_actual, value=fecha_actual.day)
        ws.cell(row=2, column=col_actual).border = thin_border
        ws.cell(row=2, column=col_actual).alignment = center_align
        ws.cell(row=2, column=col_actual).font = Font(size=9)
        ws.column_dimensions[get_column_letter(col_actual)].width = 3.5
        
        # Guardar mapeo de fecha a columna
        fecha_columna[fecha_actual] = col_actual
        
        # Manejo de encabezado de mes
        if mes_actual is None:
            mes_actual = fecha_actual.month
            col_inicio_mes = col_actual
        elif fecha_actual.month != mes_actual:
            # Guardar info del mes anterior
            columnas_por_mes.append((col_inicio_mes, col_actual - 1, mes_actual))
            mes_actual = fecha_actual.month
            col_inicio_mes = col_actual
        
        fecha_actual += timedelta(days=7)  # Avanzar una semana
        col_actual += 1
    
    # Guardar el último mes
    if mes_actual:
        columnas_por_mes.append((col_inicio_mes, col_actual - 1, mes_actual))
    
    # Ahora escribir los encabezados de mes (después de calcular todas las columnas)
    for col_ini, col_fin, mes in columnas_por_mes:
        nombre_mes = meses_espanol[mes]
        
        # Primero hacer el merge si hay más de una columna
        if col_fin > col_ini:
            ws.merge_cells(
                start_row=1, start_column=col_ini,
                end_row=1, end_column=col_fin
            )
        
        # Luego escribir el valor en la celda principal
        ws.cell(row=1, column=col_ini, value=nombre_mes)
        ws.cell(row=1, column=col_ini).font = header_font
        ws.cell(row=1, column=col_ini).alignment = center_align
        ws.cell(row=1, column=col_ini).border = thin_border
    
    return fecha_columna


def marcar_periodos_gantt(ws, fila, periodos, fecha_columna):
    """
    Marca las celdas correspondientes a los períodos con 'x'.
    - 'x' negra para estados normales
    - 'x' roja para Completada o Terminada
    No modifica los bordes (deben aplicarse antes de llamar esta función).
    """
    center_align = Alignment(horizontal='center', vertical='center')
    
    for fecha_inicio, fecha_fin, estado in periodos:
        # Determinar color de la x según estado
        es_completado = estado in [EstadoActividad.COMPLETADA, EstadoActividad.TERMINADA]
        color_x = 'FF0000' if es_completado else '000000'  # Rojo o Negro
        font_x = Font(color=color_x, bold=True)
        
        for fecha, columna in fecha_columna.items():
            # Verificar si la semana intersecta con el período
            fin_semana = fecha + timedelta(days=6)
            if not (fecha_fin < fecha or fecha_inicio > fin_semana):
                # Hay intersección - escribir x
                celda = ws.cell(row=fila, column=columna)
                celda.value = 'x'
                celda.font = font_x
                celda.alignment = center_align


def copiar_estilo_celda(celda_origen, celda_destino):
    """
    Copia el estilo de una celda a otra.
    """
    if celda_origen.has_style:
        celda_destino.font = copy(celda_origen.font)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.number_format = copy(celda_origen.number_format)
        celda_destino.protection = copy(celda_origen.protection)
        celda_destino.alignment = copy(celda_origen.alignment)


def copiar_fila_estilo(ws, fila_origen, fila_destino, max_col=100):
    """
    Copia los estilos de una fila a otra.
    """
    for col in range(1, max_col + 1):
        celda_origen = ws.cell(row=fila_origen, column=col)
        celda_destino = ws.cell(row=fila_destino, column=col)
        copiar_estilo_celda(celda_origen, celda_destino)


def calcular_altura_fila(ws, fila, columnas_a_revisar, altura_linea=15):
    """
    Calcula la altura necesaria para una fila basándose en el contenido de las celdas.
    Considera saltos de línea y el ancho de columna para texto largo.
    
    Args:
        ws: Worksheet
        fila: Número de fila
        columnas_a_revisar: Lista de columnas a revisar (ej: [4, 5, 6])
        altura_linea: Altura base por línea de texto (default 15)
    
    Returns:
        Altura calculada para la fila
    """
    max_lineas = 1
    
    for col in columnas_a_revisar:
        celda = ws.cell(row=fila, column=col)
        valor = celda.value
        
        if valor is None:
            continue
            
        texto = str(valor)
        
        # Contar líneas explícitas (saltos de línea)
        lineas_explicitas = texto.count('\n') + 1
        
        # Estimar líneas adicionales por texto largo (wrap)
        ancho_columna = ws.column_dimensions[get_column_letter(col)].width or 10
        caracteres_por_linea = max(1, int(ancho_columna * 1.2))  # Aproximación
        
        lineas_por_wrap = 0
        for linea in texto.split('\n'):
            if len(linea) > caracteres_por_linea:
                lineas_por_wrap += (len(linea) // caracteres_por_linea)
        
        total_lineas = lineas_explicitas + lineas_por_wrap
        max_lineas = max(max_lineas, total_lineas)
    
    return max(15, max_lineas * altura_linea)


def aplicar_altura_filas(ws, fila_inicio, fila_fin, columnas_a_revisar):
    """
    Aplica altura automática a un rango de filas.
    """
    for fila in range(fila_inicio, fila_fin + 1):
        altura = calcular_altura_fila(ws, fila, columnas_a_revisar)
        ws.row_dimensions[fila].height = altura


def escribir_celda_merged(ws, fila, col_inicio, col_fin, valor, thin_border, wrap_align):
    """
    Escribe un valor en celdas que pueden necesitar merge.
    """
    ws.cell(row=fila, column=col_inicio, value=valor)
    ws.cell(row=fila, column=col_inicio).border = thin_border
    ws.cell(row=fila, column=col_inicio).alignment = wrap_align
    if col_fin > col_inicio:
        ws.merge_cells(
            start_row=fila, start_column=col_inicio,
            end_row=fila, end_column=col_fin
        )


def exportar_gantt_excel(proyecto_id):
    """
    Genera un archivo Excel con la carta Gantt del proyecto.
    Retorna un BytesIO con el archivo Excel.
    Mantiene el formato original del archivo Vacio.xlsx.
    """
    # Obtener datos
    proyecto, actividades, actividades_difusion = obtener_datos_proyecto(proyecto_id)
    
    # Cargar plantilla vacía
    ruta_plantilla = os.path.join(settings.BASE_DIR, 'frontend', 'static', 'Vacio.xlsx')
    wb = load_workbook(ruta_plantilla)
    ws = wb.active
    
    # Estructura del archivo Vacio.xlsx:
    # Fila 1: Meses (celdas mergeadas)
    # Fila 2: Encabezados actividades normales
    # Fila 3: Plantilla fila actividad normal
    # Fila 4: "Difusión" (separador)
    # Fila 5: Encabezados difusión
    # Fila 6: Plantilla fila difusión
    
    FILA_PLANTILLA_NORMAL = 3
    FILA_DIFUSION_ORIGINAL = 4
    FILA_ENCABEZADO_DIFUSION = 5
    FILA_PLANTILLA_DIFUSION = 6
    
    # Calcular cuántas filas normales necesitamos insertar
    num_actividades = len(actividades)
    filas_a_insertar_normales = max(0, num_actividades - 1)  # Ya hay 1 fila plantilla
    
    # IMPORTANTE: Deshacer TODOS los merges de columnas A-B en filas de datos ANTES de insertar
    # Esto evita que los merges de la plantilla se expandan automáticamente
    merges_a_eliminar = []
    for merge in list(ws.merged_cells.ranges):
        # Eliminar merges en columnas A o B que estén en filas >= 3
        if merge.min_row >= FILA_PLANTILLA_NORMAL and merge.min_col <= 2:
            merges_a_eliminar.append(str(merge))
    for merge_str in merges_a_eliminar:
        ws.unmerge_cells(merge_str)
    
    # Insertar filas para actividades normales (antes de Difusión)
    if filas_a_insertar_normales > 0:
        ws.insert_rows(FILA_PLANTILLA_NORMAL + 1, filas_a_insertar_normales)
        # Copiar estilos de la fila plantilla a las nuevas filas
        for i in range(filas_a_insertar_normales):
            copiar_fila_estilo(ws, FILA_PLANTILLA_NORMAL, FILA_PLANTILLA_NORMAL + 1 + i)
    
    # Actualizar posiciones después de inserción
    fila_difusion = FILA_DIFUSION_ORIGINAL + filas_a_insertar_normales
    fila_encabezado_dif = FILA_ENCABEZADO_DIFUSION + filas_a_insertar_normales
    fila_plantilla_dif = FILA_PLANTILLA_DIFUSION + filas_a_insertar_normales
    
    # Calcular filas de difusión a insertar
    num_actividades_dif = len(actividades_difusion)
    filas_a_insertar_difusion = max(0, num_actividades_dif - 1)
    
    # Insertar filas para actividades de difusión
    if filas_a_insertar_difusion > 0:
        ws.insert_rows(fila_plantilla_dif + 1, filas_a_insertar_difusion)
        for i in range(filas_a_insertar_difusion):
            copiar_fila_estilo(ws, fila_plantilla_dif, fila_plantilla_dif + 1 + i)
    
    # Calcular rango de fechas
    fecha_min, fecha_max = obtener_rango_fechas(actividades, actividades_difusion)
    
    # Deshacer los merges de la fila 1 (meses) para poder regenerarlos
    merges_fila1 = [m for m in list(ws.merged_cells.ranges) if m.min_row == 1 and m.min_col >= 7]
    for merge in merges_fila1:
        ws.unmerge_cells(str(merge))
    
    # Generar columnas de fechas (columna G = 7)
    fecha_columna = generar_columnas_fechas(ws, fecha_min, fecha_max, col_inicio=7)
    
    # Escribir actividades normales con merge vertical por línea de trabajo
    fila_actual = FILA_PLANTILLA_NORMAL
    id_incremental = 1
    
    # Agrupar actividades por línea de trabajo consecutiva
    grupos_linea = []
    if actividades:
        grupo_actual = {'linea': actividades[0]['linea_trabajo'], 'actividades': [actividades[0]], 'fila_inicio': fila_actual}
        for act in actividades[1:]:
            if act['linea_trabajo'] == grupo_actual['linea']:
                grupo_actual['actividades'].append(act)
            else:
                grupos_linea.append(grupo_actual)
                grupo_actual = {'linea': act['linea_trabajo'], 'actividades': [act], 'fila_inicio': fila_actual + len(grupos_linea[-1]['actividades']) if grupos_linea else fila_actual}
        grupos_linea.append(grupo_actual)
    
    # Recalcular filas de inicio para cada grupo
    fila_temp = FILA_PLANTILLA_NORMAL
    for grupo in grupos_linea:
        grupo['fila_inicio'] = fila_temp
        fila_temp += len(grupo['actividades'])
    
    # Estilos para bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_bottom_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='medium')  # Borde inferior grueso
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Borde especial para la celda mergeada de línea de trabajo
    linea_border_normal = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    linea_border_bottom = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='medium')
    )
    
    for grupo in grupos_linea:
        fila_inicio_grupo = grupo['fila_inicio']
        fila_fin_grupo = fila_inicio_grupo + len(grupo['actividades']) - 1
        
        # Deshacer merges existentes en el rango de la línea de trabajo
        for fila in range(fila_inicio_grupo, fila_fin_grupo + 1):
            for merge in list(ws.merged_cells.ranges):
                if merge.min_row == fila and merge.min_col == 1:
                    ws.unmerge_cells(str(merge))
                    break
        
        # Merge vertical de línea de trabajo (columnas A-B)
        ws.merge_cells(
            start_row=fila_inicio_grupo, start_column=1,
            end_row=fila_fin_grupo, end_column=2
        )
        
        # Escribir valor y aplicar estilo a la celda principal del merge
        celda_linea = ws.cell(row=fila_inicio_grupo, column=1, value=grupo['linea'])
        celda_linea.alignment = center_align
        celda_linea.border = linea_border_normal
        
        # Aplicar bordes a todas las celdas del rango mergeado para que se vean
        for fila in range(fila_inicio_grupo, fila_fin_grupo + 1):
            for col in range(1, 3):
                celda = ws.cell(row=fila, column=col)
                if fila == fila_fin_grupo:
                    celda.border = linea_border_bottom
                else:
                    celda.border = linea_border_normal
        
        # Escribir cada actividad del grupo
        for idx, act in enumerate(grupo['actividades']):
            fila = fila_inicio_grupo + idx
            es_ultima_fila_grupo = (idx == len(grupo['actividades']) - 1)
            
            # Determinar borde a usar (grueso en última fila del grupo)
            borde_actual = thick_bottom_border if es_ultima_fila_grupo else thin_border
            
            # Alignment con wrap_text para las celdas de texto
            wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Columna C: N° (ID incremental)
            ws.cell(row=fila, column=3, value=id_incremental)
            ws.cell(row=fila, column=3).border = borde_actual
            ws.cell(row=fila, column=3).alignment = Alignment(horizontal='center', vertical='center')
            
            # Columna D: Nombre actividad
            ws.cell(row=fila, column=4, value=act['nombre'])
            ws.cell(row=fila, column=4).border = borde_actual
            ws.cell(row=fila, column=4).alignment = wrap_align
            
            # Columna E: Responsables
            ws.cell(row=fila, column=5, value='; '.join(act['responsables']))
            ws.cell(row=fila, column=5).border = borde_actual
            ws.cell(row=fila, column=5).alignment = wrap_align
            
            # Columna F: Producto asociado
            ws.cell(row=fila, column=6, value=act['producto'])
            ws.cell(row=fila, column=6).border = borde_actual
            ws.cell(row=fila, column=6).alignment = wrap_align
            
            # Aplicar borde a TODAS las celdas de fecha de esta fila primero
            for columna in fecha_columna.values():
                ws.cell(row=fila, column=columna).border = borde_actual
            
            # Marcar períodos con x (sin modificar bordes, ya están aplicados)
            marcar_periodos_gantt(ws, fila, act['periodos'], fecha_columna)
            
            id_incremental += 1
        
        fila_actual = fila_fin_grupo + 1
    
    # Combinar celdas del título "Difusión" para que cubra toda la Gantt
    # Calcular la última columna de fechas
    ultima_columna_fecha = max(fecha_columna.values()) if fecha_columna else 6
    
    # Deshacer merges existentes en la fila de difusión
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row == fila_difusion:
            ws.unmerge_cells(str(merge))
    
    # Hacer merge del título "Difusión" desde columna A hasta la última columna de fechas
    ws.merge_cells(
        start_row=fila_difusion, start_column=1,
        end_row=fila_difusion, end_column=ultima_columna_fecha
    )
    ws.cell(row=fila_difusion, column=1, value="Difusión")
    ws.cell(row=fila_difusion, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=fila_difusion, column=1).font = Font(bold=True, size=11, color='FFFFFF')
    
    # Combinar celdas B-C del encabezado de difusión (fila_encabezado_dif)
    # Deshacer merges existentes en el encabezado
    for merge in list(ws.merged_cells.ranges):
        if merge.min_row == fila_encabezado_dif and merge.min_col == 2:
            ws.unmerge_cells(str(merge))
            break
    
    # Hacer merge de B-C en el encabezado
    ws.merge_cells(
        start_row=fila_encabezado_dif, start_column=2,
        end_row=fila_encabezado_dif, end_column=3
    )
    
    # Escribir actividades de difusión
    fila_actual = fila_plantilla_dif
    id_incremental = 1
    
    for act in actividades_difusion:
        # Deshacer merge de B-C si existe
        for merge in list(ws.merged_cells.ranges):
            if merge.min_row == fila_actual and merge.min_col == 2:
                ws.unmerge_cells(str(merge))
                break
        
        # Columna A: N° (ID incremental)
        ws.cell(row=fila_actual, column=1, value=id_incremental)
        ws.cell(row=fila_actual, column=1).border = thin_border
        
        # Columnas B-C: Nombre actividad
        ws.cell(row=fila_actual, column=2, value=act['nombre'])
        ws.cell(row=fila_actual, column=2).border = thin_border
        ws.cell(row=fila_actual, column=3).border = thin_border
        ws.merge_cells(start_row=fila_actual, start_column=2, end_row=fila_actual, end_column=3)
        
        # Columna D: Responsables
        ws.cell(row=fila_actual, column=4, value='; '.join(act['responsables']))
        ws.cell(row=fila_actual, column=4).border = thin_border
        
        # Columna E: Productos
        ws.cell(row=fila_actual, column=5, value='; '.join(act['productos']))
        ws.cell(row=fila_actual, column=5).border = thin_border
        
        # Columna F: Líneas de trabajo
        ws.cell(row=fila_actual, column=6, value='; '.join(act['lineas_trabajo']))
        ws.cell(row=fila_actual, column=6).border = thin_border
        
        # Aplicar borde a todas las celdas de fecha
        for columna in fecha_columna.values():
            ws.cell(row=fila_actual, column=columna).border = thin_border
        
        # Marcar períodos con x
        marcar_periodos_gantt(ws, fila_actual, act['periodos'], fecha_columna)
        
        fila_actual += 1
        id_incremental += 1
    
    # Limpiar filas sobrantes si no hay actividades
    if num_actividades == 0:
        # Limpiar fila 3 plantilla
        for col in range(1, 7):
            ws.cell(row=FILA_PLANTILLA_NORMAL, column=col, value="")
    
    if num_actividades_dif == 0:
        # Limpiar fila plantilla difusión
        for col in range(1, 7):
            ws.cell(row=fila_plantilla_dif, column=col, value="")
    
    # Aplicar altura automática a las filas de actividades normales
    # Columnas a revisar: D (nombre), E (responsables), F (producto)
    if num_actividades > 0:
        fila_fin_normales = FILA_PLANTILLA_NORMAL + num_actividades - 1
        aplicar_altura_filas(ws, FILA_PLANTILLA_NORMAL, fila_fin_normales, [4, 5, 6])
    
    # Aplicar altura automática a las filas de actividades de difusión
    # Columnas a revisar: B-C (nombre), D (responsables), E (productos), F (líneas)
    if num_actividades_dif > 0:
        fila_fin_difusion = fila_plantilla_dif + num_actividades_dif - 1
        aplicar_altura_filas(ws, fila_plantilla_dif, fila_fin_difusion, [2, 4, 5, 6])
    
    # Guardar en BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output, f"Gantt_{proyecto.nombre}.xlsx"
